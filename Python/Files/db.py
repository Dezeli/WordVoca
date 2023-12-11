import sqlite3
import datetime
from gtts import gTTS
from openpyxl import load_workbook, Workbook


def add_db(lang, word, content):
    db = sqlite3.connect("../../Database/Words.db")
    cursor = db.cursor()
    cursor.execute(f"SELECT COUNT(*) FROM Words")
    count = cursor.fetchone()[0]
    date = datetime.datetime.now()
    insert_query = f'INSERT INTO Words VALUES("{count+1}", "{lang}", "{word}", "{content}", "{date}", "False")'
    cursor.execute(insert_query)
    db.commit()
    tts1 = gTTS(text=word, lang=lang)
    tts1.save(f"../../Sound/{count+1}_1.mp3")
    tts2 = gTTS(text=content, lang="ko")
    tts2.save(f"../../Sound/{count+1}_2.mp3")


def get_ran_word(lang):
    db = sqlite3.connect("../../Database/Words.db")
    cursor = db.cursor()
    cursor.execute(
        f'SELECT Word, Content, Seq FROM Words Where Del=="False" and Type=="{lang}" Order by random()'
    )
    return cursor.fetchall()


def get_word_num():
    db = sqlite3.connect("../../Database/Words.db")
    cursor = db.cursor()
    cursor.execute(f'SELECT Count(*) FROM Words Where Del=="False"')
    return cursor.fetchone()[0]


def get_all_words(sort):
    db = sqlite3.connect("../../Database/Words.db")
    cursor = db.cursor()
    if sort[1] == True:
        cursor.execute(
            f'SELECT Seq, Type, Word, Content, Date FROM Words Where Del=="False" ORDER BY {sort[0]}'
        )
    else:
        cursor.execute(
            f'SELECT Seq, Type, Word, Content, Date FROM Words Where Del=="False" ORDER BY {sort[0]} DESC'
        )
    return cursor.fetchall()


def delete_word(seq):
    db = sqlite3.connect("../../Database/Words.db")
    cursor = db.cursor()
    cursor.execute(f'UPDATE Words SET Del="True" Where Seq=={seq}')
    db.commit()


def add_excel_file():
    load_wb = load_workbook("../../어휘추가.xlsx", data_only=True)
    load_ws = load_wb["Sheet1"]
    num = 4
    excel_li = []
    error = False
    while True:
        num += 1
        word = load_ws[f"B{num}"].value
        content = load_ws[f"C{num}"].value
        lang = load_ws[f"D{num}"].value
        if word and content:
            if lang == "ENG":
                add_db("en", word, content)
            elif lang == "KOR":
                add_db("ko", word, content)
            excel_li.append([num, word, content, lang])
        else:
            # for i in range(2,5):
            #     load_ws.cell(row=num, column=i).value = ''
            #     load_wb.save("../../어휘추가.xlsx")
            break
        try:
            for i in range(2, 5):
                load_ws.cell(row=num, column=i).value = ""
            load_wb.save("../../어휘추가.xlsx")
        except:
            error = True

    return excel_li[-1], error


if __name__ == "__main__":
    add_excel_file()
